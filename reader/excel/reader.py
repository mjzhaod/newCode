import openpyxl

# 入参定义方式：QueryInfoHolder类型，定义的是你要查询一行中的第几列，比如想查询第二列就是 1
# 查询数据思想：根据查询字段将excel中的一行数据封装为一个ResultSet，
# 封装结果思想：现默认查询出来为一个数组，数组中封装方式由result_handler回调函数定义，回调函数的参数为resultSet，可根据resultSet自行封装为任意对象
# 缺点： 无法按列读取数据


def read(file_reader_holder):
    """
    思想为：读取excel中行，根据传入查询字段列表封装为result_set, 随后使用回调函数封装用户自定结构

    :param file_reader_holder: FileReaderInfoHolder
    :return: 数组，数组中为结果封装回调函数返回的类型
    """

    workbook = openpyxl.load_workbook(file_reader_holder.path)
    sheet_name = file_reader_holder.sheet_name
    if sheet_name is None or sheet_name == "":
        sheet_name = workbook.sheetnames[0]
    sheet = workbook[sheet_name]
    # TODO check property 's index is out of range

    properties = file_reader_holder.properties
    results = []

    for row in sheet.iter_rows():
        result_set = row_to_result_set(row, properties)
        # 调用回调函数组装结果 invoke result handler to build result
        result = file_reader_holder.result_handler(result_set)
        # 将结果添加到数组中 collect result to an array
        results.append(result)
    return results


def row_to_result_set(row, properties):
    """
    根据定义的查询字段获得结果集
    :param row:  excel中遍历行
    :param properties: 查询结构中定义的查询列
    :return: 返回结果集 result_set
    """
    results = []
    for prop in properties:
        value = row[prop.index].value
        result = Result(prop.name, value)
        results.append(result)
    return ResultSet(results)


class ResultSet:

    result = []

    def __init__(self, result):
        self.result = result

    # 根据位置索引获取结果集中的值
    def get_by_index(self, index):
        return self.result[index].value

    def get_by_name(self, name):
        for cell in self.result:
            if cell.name == name:
                return cell.value


class Result:
    name = "",
    value = object

    def __init__(self, name, value):
        self.name = name
        self.value = value


class FileReaderInfoHolder:
    """
        定义文件读取信息

        Parameters

        properties: Property
            表示待查询列信息
        path: str
            表示待读取文件位置
        sheet_name: str
            待读取sheet 表名称
        result_handler: result_handler
            处理结果集回调函数，参数为result_set

    """

    properties = [],

    path = "",

    sheet_name = "",

    result_handler = object,

    def __init__(self):
        pass

    def __init__(self, properties, path, sheet_name, result_handler):
        self.properties = properties
        self.path = path
        self.sheet_name = sheet_name
        self.result_handler = result_handler


class QueryInfoHolder:

    # 当前属性在哪一列
    index = 0
    name = ''

    def __init__(self, index, name=''):
        self.index = index
        self.name = name


class WriteInfoHolder:

    index = 0

    def __init__(self, index):
        self.index = index

