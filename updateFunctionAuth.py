import requests
import json
import openpyxl
from request import Request
from do_request import do_request,build_url,check
import data_base


def generate_result(row_num, columns):
    re = Request()
    for column in columns:
        value = sheet.cell(row_num, column["index"]).value
        setattr(re, column["name"], value)
    return re


workbook = openpyxl.load_workbook(data_base.address)
sheet=workbook.get_sheet_by_name(r"功能权限")
urls=''
headers=''
request_parameter=''

for rownum in range(2,sheet.max_row+1):
    # 定义excel里的一列 对应request对象里的字段 name表示request里的属性名， index表示excel的列号
    request = generate_result(rownum, [{"name": "url", "index": 3}, {"name": "headers", "index": 4},
                                         {"name": "method", "index": 5},
                                         {"name": "body", "index": 6}],{"name":"expectResult","index":7})
    build_url(request, data_base.token)
    response=do_request(request)
    result=check(request,response)
    if result:
        print()



    print(response.status_code)
    print(response.text)


