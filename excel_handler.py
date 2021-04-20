import openpyxl




def read_excel(file_name,name,column):

    data=openpyxl.load_workbook(file_name)
    sheet_name=data.sheetnames
    table=data[sheet_name[0]]
    for rownum in range(2,sheet.max_row+1):





def write(file_name):
    temp_result=read(file_name)
    wb = openpyxl.load_workbook(file_name)
    sheet_name = wb.sheetnames
    table = wb[sheet_name[0]]
    index = 1
    column=7

    for i in table.iter_rows():
        if index == 1:
            pass
        else:
            table.cell(row=index, column=column).value =temp_result[index-2]
        index += 1
    wb.save(file_name)
